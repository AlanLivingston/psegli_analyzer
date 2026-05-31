#!/usr/bin/env python3
"""
PSEG Long Island consumption analyzer.

Reads a downloaded interval-consumption CSV (consumption meter + optional
generation meter whose number ends in 'g'), splits it into billing cycles, and
writes an .xlsx dashboard estimating the full bill - every line item that
appears on a real PSEG bill - under four rate plans: 580, 180, 194 and 195.

Rates are DATE-MATCHED: a monthly rate schedule (delivery, supply and riders)
is kept in a local cache and shown on a "Rates" tab. For each billing cycle the
applicable rate is computed by day-weighting the schedule across the cycle's
days, so a cycle that straddles a rate change (e.g. the Jan 1 increase) or a
season boundary gets the correct blended rate - the way PSEG actually bills.

Historical rates are available on PSEG LI back to June 2025:
  https://www.psegliny.com/aboutpseglongisland/ratesandtariffs/rateinformation
The scraper refreshes/extends the cache from that page when the data covers
months not yet downloaded. Dates before June 2025 fall back to the earliest
known rates and a warning is printed.

Usage:
    python pseg_analyzer.py --cons_file Usage.csv
    python pseg_analyzer.py --cons_file Usage.csv --read-dates 2025-11-15 2025-12-15 2026-01-16 ...
    python pseg_analyzer.py --cons_file Usage.csv --demand 16 --out report.xlsx [--refresh]
"""

import argparse
import copy
import datetime as dt
import json
import math
import os
import re
import sys
from typing import Any, cast

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CACHE_FILE = "pseg_rates_cache.json"
RATE_INFO_URL = "https://www.psegliny.com/aboutpseglongisland/ratesandtariffs/rateinformation"
HIST_FLOOR = (2025, 6)   # PSEG publishes historical rates back to June 2025

# --------------------------------------------------------------------------- #
# Rate epochs. PSEG raised residential delivery/rider rates on 2026-01-01.     #
# Values confirmed from residential bills + the PSEG rate guide. Where a value #
# is not independently confirmed it is flagged (see month flags) and the       #
# scraper is expected to correct it. Seasons: Summer = Jun-Sep, Winter = rest. #
# --------------------------------------------------------------------------- #
EPOCH_PRE: dict[str, Any] = {            # in effect through 2025-12-31
    "basic_day": 0.5400, "der_rate": 0.005554, "mfc_rate": 0.001886,
    "cbc_rate": 0.0330, "dsa_rate": -0.0006, "rda_rate": -0.0003,
    "nysa_pct": 0.0027, "pilots_pct": 0.0215, "suffolk_pct": 0.0231, "salestax_pct": 0.025,
    "delivery": {
        "580": {"t1_S": 0.1064, "t1_W": 0.1021, "t2_S": 0.1348, "t2_W": 0.1021,
                "exc_S": 0.1348, "exc_W": 0.0561, "psc_mult": 1.00},
        "180": {"t1_S": 0.1064, "t1_W": 0.1021, "exc_S": 0.1348, "exc_W": 0.1021, "psc_mult": 1.00},
        "194": {"off_S": 0.1093, "off_W": 0.0929, "peak_S": 0.2217, "peak_W": 0.1885,
                "mult_off": 0.83, "mult_peak_S": 1.9505, "mult_peak_W": 2.1340},
        "195": {"so_S": 0.0452, "so_W": 0.0450, "off_S": 0.1388, "off_W": 0.0929,
                "peak_S": 0.2979, "peak_W": 0.2440, "mult_so": 0.60, "mult_off": 1.00,
                "mult_peak_S": 1.7419, "mult_peak_W": 1.9688},
    },
}
EPOCH_2026: dict[str, Any] = {           # in effect from 2026-01-01
    "basic_day": 0.5600, "der_rate": 0.006014, "mfc_rate": 0.001763,
    "cbc_rate": 0.0372, "dsa_rate": -0.0006, "rda_rate": -0.0003,
    "nysa_pct": 0.0027, "pilots_pct": 0.0215, "suffolk_pct": 0.0231, "salestax_pct": 0.025,
    "delivery": {
        "580": {"t1_S": 0.1064, "t1_W": 0.1064, "t2_S": 0.1348, "t2_W": 0.1064,
                "exc_S": 0.1348, "exc_W": 0.0585, "psc_mult": 1.00},
        "180": {"t1_S": 0.1064, "t1_W": 0.1064, "exc_S": 0.1348, "exc_W": 0.1064, "psc_mult": 1.00},
        "194": {"off_S": 0.1093, "off_W": 0.0929, "peak_S": 0.2217, "peak_W": 0.1885,
                "mult_off": 0.83, "mult_peak_S": 1.9505, "mult_peak_W": 2.1340},
        "195": {"so_S": 0.0452, "so_W": 0.0450, "off_S": 0.1388, "off_W": 0.0929,
                "peak_S": 0.2979, "peak_W": 0.2440, "mult_so": 0.60, "mult_off": 1.00,
                "mult_peak_S": 1.7419, "mult_peak_W": 1.9688},
    },
}
# Monthly Power Supply Charge ($/kWh) - the time-varying supply base rate.
SEED_PSC = {
    "2025-05": 0.122603, "2025-06": 0.130053, "2025-07": 0.118428, "2025-08": 0.127798,
    "2025-09": 0.130177, "2025-10": 0.129685, "2025-11": 0.127804, "2025-12": 0.125526,
    "2026-01": 0.129777, "2026-02": 0.146948, "2026-03": 0.165666, "2026-04": 0.165927,
    "2026-05": 0.160182,
}
# Delivery/rider values are independently CONFIRMED for these months (from bills).
CONFIRMED = {"2025-10", "2025-11", "2025-12", "2026-01", "2026-02", "2026-03", "2026-04", "2026-05"}

# Local sales-tax fraction on residential electricity, by PSEG LI service area.
# Suffolk is bill-verified; Nassau exempts residential energy (except the Glen Cove
# and Long Beach school districts, ~3%); the Rockaways fall under NYC's local rate.
# The supply/delivery schedules and the PILOT/property-tax recovery are LIPA-wide and
# do not vary by county; only this sales-tax line does. Override with --sales-tax.
COUNTY_SALES_TAX = {"suffolk": 0.025, "nassau": 0.000, "rockaways": 0.045}


def apply_county_taxes(sched: "dict[str, Any]", county: str, override: "float | None") -> float:
    """Set the local sales-tax fraction on every month record for the chosen service
    area (in memory only; not persisted to the shared cache). Returns the rate used."""
    st = override if override is not None else COUNTY_SALES_TAX[county]
    for rec in sched.values():
        rec["salestax_pct"] = st
    return st

RATE_TITLES = {
    "580": "Rate 580 - Residential, Home Heating",
    "180": "Rate 180 - Residential (standard)",
    "194": "Rate 194 - Time-of-Day, Off-Peak",
    "195": "Rate 195 - Time-of-Day, Super Off-Peak",
}


# --------------------------------------------------------------------------- #
# Rate schedule (monthly) construction + cache + scraping                      #
# --------------------------------------------------------------------------- #
def month_key(y, mo):
    """Canonical 'YYYY-MM' key used throughout the rate schedule and PSC tables."""
    return f"{y}-{mo:02d}"


def build_seed_schedule():
    """Build the baseline monthly rate schedule from the hard-coded epoch tables.

    Each month gets a deep copy of the epoch in effect for it (pre-2026 vs 2026+),
    stamped with that month's Power Supply Charge and a confidence flag the rest of
    the program uses to warn the user."""
    sched = {}
    for mk, psc in SEED_PSC.items():
        y, mo = int(mk[:4]), int(mk[5:7])
        # The tariff changed on 2026-01-01, so pick the matching epoch's delivery/riders.
        epoch = EPOCH_2026 if (y, mo) >= (2026, 1) else EPOCH_PRE
        rec = copy.deepcopy(epoch)            # deep copy so per-month edits never alias the epoch
        rec["psc"] = psc
        # Confidence flag drives the console warnings and the yellow cells on the Rates tab.
        if (y, mo) < HIST_FLOOR:
            rec["flag"] = "stale"            # before published history
        elif mk in CONFIRMED:
            rec["flag"] = "confirmed"        # delivery/riders verified against a real bill
        else:
            rec["flag"] = "estimate"         # PSC exact; delivery/riders estimated
        sched[mk] = rec
    return sched


def load_cache():
    """Return the rate cache, merging any on-disk scraped values over the seed.

    On-disk months win (they may carry scraped/confirmed rates), but every seed
    month is backfilled so the schedule always covers the known range even if the
    cache file is partial. A corrupt cache degrades gracefully to the pure seed."""
    seed = build_seed_schedule()
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE) as f:
                disk = json.load(f)
            sched = disk.get("schedule", {})
            for mk, rec in seed.items():        # backfill seed months not on disk
                sched.setdefault(mk, rec)
            return {"scraped_at": disk.get("scraped_at"), "schedule": sched}
        except Exception as e:
            # Never let a bad cache file stop an analysis; fall back to the seed.
            print(f"  (cache unreadable: {e}; using seed schedule)")
    return {"scraped_at": None, "schedule": seed}


def save_cache(cache):
    """Persist the rate cache to disk as pretty-printed JSON."""
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2)


def months_in_span(a, b):
    """List every 'YYYY-MM' key from date a through date b inclusive, in order."""
    out, y, mo = [], a.year, a.month
    while (y, mo) <= (b.year, b.month):
        out.append(month_key(y, mo))
        mo += 1
        if mo > 12:                           # roll over December -> January of next year
            mo, y = 1, y + 1
    return out


def scrape_rate_information(month_keys):
    """Best-effort scrape of the PSEG LI rate-information page. Returns
    {month_key: {'psc':..., optionally delivery/rider fields}}. Network or
    parse failures degrade gracefully to {} (callers keep the seed)."""
    found = {}
    try:
        import urllib.request
        req = urllib.request.Request(RATE_INFO_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=25) as r:
            html = r.read().decode("utf-8", "ignore")
        names = ["january", "february", "march", "april", "may", "june", "july",
                 "august", "september", "october", "november", "december"]
        for i, nm in enumerate(names, 1):
            for mt in re.finditer(rf"{nm}\s+(\d{{4}}).{{0,120}}?\$?\s*(0\.\d{{5,6}})",
                                  html, re.IGNORECASE | re.DOTALL):
                mk = month_key(int(mt.group(1)), i)
                if mk in month_keys:
                    found.setdefault(mk, {})["psc"] = float(mt.group(2))
    except Exception as e:
        print(f"  (scrape failed: {e})")
    return found


def refresh_schedule(cache, start, end, force=False):
    """Ensure the schedule covers [start, end], scraping any missing months.

    Normally only months absent from the cache are fetched; with force=True the
    months still flagged 'estimate' are re-fetched too. Anything that can't be
    scraped is carried forward from the latest known month and flagged 'carried'.
    The cache is saved (with a timestamp) only when a fetch actually happened."""
    sched = cache["schedule"]
    needed = months_in_span(start, end)
    missing = [m for m in needed if m not in sched]
    # Targets = months we don't have, plus (when forcing) months whose rates are still estimated.
    target = set(missing) | ({m for m in needed if sched.get(m, {}).get("flag") == "estimate"}
                             if force else set())
    if not target:
        print("  Rate schedule already covers the data period; no scrape needed.")
        return cache
    print(f"  Fetching rates for {len(target)} month(s) from PSEG LI rate-information...")
    got = scrape_rate_information(set(target))
    for mk, fields in got.items():
        # Overlay scraped fields onto the existing record (or a fresh 2026 epoch if brand new).
        base = sched.get(mk) or copy.deepcopy(EPOCH_2026)
        base.update({k: v for k, v in fields.items()})
        base["flag"] = "scraped"
        sched[mk] = base
    # Any month we needed but still couldn't get: carry the most recent record forward.
    still = [m for m in missing if m not in sched]
    if still:
        last = sorted(sched)[-1]
        for m in still:
            rec = copy.deepcopy(sched[last]); rec["flag"] = "carried"
            sched[m] = rec
        print(f"  Could not obtain {', '.join(still)}; carried forward {last}. Verify on psegliny.com.")
    cache["scraped_at"] = dt.datetime.now().isoformat(timespec="seconds")
    save_cache(cache)
    return cache


def get_month(sched, y, mo, warnings):
    """Return the monthly rate record for (y,mo), with fallback + warning."""
    mk = month_key(y, mo)
    if mk in sched:
        rec = sched[mk]
        if rec.get("flag") == "stale":
            warnings.add(f"Dates in/before {mk} predate published history (back to "
                         f"{HIST_FLOOR[0]}-{HIST_FLOOR[1]:02d}); using earliest known rates.")
        elif rec.get("flag") in ("estimate", "carried"):
            warnings.add(f"Delivery/rider rates for {mk} are estimated (PSC may be exact); "
                         f"run with --refresh once online to scrape exact values.")
        return rec
    keys = sorted(sched)
    fb = keys[0] if (y, mo) < tuple(map(int, keys[0].split("-"))) else keys[-1]
    warnings.add(f"No rate data for {mk}; used nearest available ({fb}).")
    return sched[fb]


# --------------------------------------------------------------------------- #
# Consumption loading + usage primitives + per-cycle effective rates           #
# --------------------------------------------------------------------------- #
def load_consumption(path):
    """Parse a PSEG LI interval CSV into a per-timestamp DataFrame of imp/exp/net kWh.

    The export (generation) meter shares the consumption meter number with a 'g'
    suffix, so its rows are split out by a regex on the Meter column, made positive,
    and subtracted from imports to get net per 15-minute interval."""
    df = pd.read_csv(path)
    # Accept any capitalization of the three required columns, then normalize names.
    cols = {c.lower(): c for c in df.columns}
    if not all(n in cols for n in ("start", "meter", "kwh")):
        sys.exit(f"CSV must have Start, Meter, kWh columns; found {list(df.columns)}")
    df = df.rename(columns={cols["start"]: "Start", cols["meter"]: "Meter", cols["kwh"]: "kWh"})
    try:
        # PSEG's usual format; fall back to pandas' parser for any other layout.
        df["ts"] = pd.to_datetime(df["Start"], format="%m/%d/%Y %I:%M:%S %p")
    except (ValueError, TypeError):
        df["ts"] = pd.to_datetime(df["Start"])
    df["kWh"] = cast("pd.Series", pd.to_numeric(df["kWh"], errors="coerce")).fillna(0.0)
    # Generation rows carry a meter id like "#12345 g"; everything else is consumption.
    df["is_gen"] = df["Meter"].astype(str).str.contains(r"#\s*\d+\s*g", case=True, regex=True)
    con = cast("pd.DataFrame", df[~df["is_gen"]])
    gen = cast("pd.DataFrame", df[df["is_gen"]]).copy()
    gen["kWh"] = gen["kWh"].abs()             # exports are stored negative; make them positive
    imp = con.groupby("ts")["kWh"].sum()
    exp = gen.groupby("ts")["kWh"].sum()
    m = pd.DataFrame({"imp": imp, "exp": exp}).fillna(0.0).sort_index()
    if m.empty:
        sys.exit("No usable rows parsed from the CSV.")
    m["net"] = m["imp"] - m["exp"]            # positive = drew from grid, negative = exported
    return m


def federal_holidays(years):
    """Return the set of US federal holiday dates across the given years.

    On-peak pricing is suspended on these days, so cycle_primitives needs them to
    classify weekday afternoons correctly. Covers both fixed-date and floating
    (nth-weekday / last-Monday) holidays."""
    hol = set()
    def nth(y, mo, wd, n):
        """Date of the n-th weekday `wd` (Mon=0) in month `mo` of year `y`."""
        d = dt.date(y, mo, 1) + dt.timedelta((wd - dt.date(y, mo, 1).weekday()) % 7)
        return d + dt.timedelta(weeks=n - 1)
    def last(y, mo, wd):
        """Date of the last weekday `wd` (Mon=0) in month `mo` of year `y`."""
        d = dt.date(y, mo, 28)
        while d.month == mo:                  # walk to the last day of the month
            d += dt.timedelta(days=1)
        d -= dt.timedelta(days=1)
        return d - dt.timedelta((d.weekday() - wd) % 7)
    for y in years:
        # Fixed-date holidays: New Year, Juneteenth, July 4, Veterans Day, Christmas.
        for mo, da in [(1, 1), (6, 19), (7, 4), (11, 11), (12, 25)]:
            hol.add(dt.date(y, mo, da))
        # Floating holidays: MLK, Presidents, Memorial, Labor, Columbus, Thanksgiving.
        hol.add(nth(y, 1, 0, 3)); hol.add(nth(y, 2, 0, 3)); hol.add(last(y, 5, 0))
        hol.add(nth(y, 9, 0, 1)); hol.add(nth(y, 10, 0, 2)); hol.add(nth(y, 11, 3, 4))
    return hol


def season_of(month):
    """Season code for a month: 'S' (summer, Jun-Sep) or 'W' (winter, the rest)."""
    return "S" if month in (6, 7, 8, 9) else "W"


def effective_rates(a, b, sched, warnings):
    """Day-weight every rate parameter across the cycle [a,b], picking each
    day's month and season. Returns a flat dict of blended values."""
    keys_shared = ["basic_day", "der_rate", "mfc_rate", "cbc_rate", "dsa_rate", "rda_rate",
                   "nysa_pct", "pilots_pct", "suffolk_pct", "salestax_pct", "psc"]
    delivery_spec = {
        "580": ["t1", "t2", "exc"], "180": ["t1", "exc"],
        "194": ["off", "peak"], "195": ["so", "off", "peak"],
    }
    mult_spec = {  # output key -> (plan, field, seasonal?) for the PSC supply multipliers
        "m194_off": ("194", "mult_off", False), "m194_peak": ("194", "mult_peak", True),
        "m195_so": ("195", "mult_so", False), "m195_off": ("195", "mult_off", False),
        "m195_peak": ("195", "mult_peak", True),
    }
    # Seed an accumulator for every output key; we sum one day at a time then divide by n.
    acc = {k: 0.0 for k in keys_shared}
    for plan, fields in delivery_spec.items():
        for f in fields:
            acc[f"d{plan}_{f}"] = 0.0
    for k in mult_spec:
        acc[k] = 0.0
    n = 0
    d = a
    while d <= b:
        # Each calendar day contributes its month's rates and its season's tiers/multipliers.
        rec = get_month(sched, d.year, d.month, warnings)
        s = season_of(d.month)
        for k in keys_shared:
            acc[k] += rec[k]
        for plan, fields in delivery_spec.items():
            for f in fields:
                acc[f"d{plan}_{f}"] += rec["delivery"][plan][f"{f}_{s}"]
        for ok, (plan, field, seasonal) in mult_spec.items():
            key = f"{field}_{s}" if seasonal else field   # only peak multipliers vary by season
            acc[ok] += rec["delivery"][plan][key]
        n += 1
        d += dt.timedelta(days=1)
    return {k: v / n for k, v in acc.items()}   # day-weighted average over the cycle


def cycle_primitives(m, cycles, holidays, sched, warnings):
    """Reduce the interval data to one summary record per billing cycle.

    For each cycle [a, b] it sums net kWh overall and within the on-peak and
    super-off-peak windows (deriving the two off-peak figures by subtraction),
    estimates peak demand, and attaches the day-weighted effective rates. The TOD
    window rules: on-peak = weekday 3-7 p.m. excluding federal holidays;
    super-off-peak = 10 p.m.-6 a.m.; off-peak = everything else."""
    idx = cast("pd.DatetimeIndex", m.index)
    dtacc = idx.to_series().dt
    hour = dtacc.hour.to_numpy()
    dow = dtacc.dayofweek.to_numpy()
    dates = dtacc.date.to_numpy()
    is_hol = np.array([d in holidays for d in dates])
    # Classify each 15-minute interval into its time-of-day bucket.
    is_peak = (dow < 5) & (~is_hol) & (hour >= 15) & (hour < 19)
    is_so = (hour >= 22) | (hour < 6)
    m = m.copy(); m["peak"] = is_peak; m["so"] = is_so
    rows = []
    for (a, b) in cycles:
        # Inclusive date mask; the caller passes b = next_read - 1 day for half-open cycles.
        sel = (dates >= a) & (dates <= b)
        g = cast("pd.DataFrame", m[sel])
        if g.empty:
            continue
        net = float(g["net"].to_numpy().sum())
        peak = float(g.loc[g["peak"], "net"].to_numpy().sum())
        so = float(g.loc[g["so"], "net"].to_numpy().sum())
        eff = effective_rates(a, b, sched, warnings)
        rows.append(dict(start=a, end=b, end_label=b, days=(b - a).days + 1, net=net, peak=peak, so=so,
                         # 195 off-peak excludes both peak and super-off; 194 off-peak is just non-peak.
                         off195=net - peak - so, off194=net - peak,
                         # Demand ~ max 15-min import x4 (kW), rounded up; override with --demand.
                         demand=math.ceil(float(g["imp"].to_numpy().max()) * 4), eff=eff))
    return rows


# --------------------------------------------------------------------------- #
# Workbook                                                                     #
# --------------------------------------------------------------------------- #
FONT = "Arial"
BLUE, BLACK, GREEN, WHITE = "0000FF", "000000", "008000", "FFFFFF"
HEAD = PatternFill("solid", fgColor="0F6E56")
SUB = PatternFill("solid", fgColor="E1F5EE")
TOT = PatternFill("solid", fgColor="F1EFE8")
YEL = PatternFill("solid", fgColor="FFF6CC")
MONEY = '$#,##0.00;($#,##0.00)'
KWH = '#,##0'
RFMT = '0.000000'
EFF_KEYS = ["basic_day", "der_rate", "mfc_rate", "cbc_rate", "dsa_rate", "rda_rate",
            "nysa_pct", "pilots_pct", "suffolk_pct", "salestax_pct", "psc",
            "d580_t1", "d580_t2", "d580_exc", "d180_t1", "d180_exc",
            "d194_off", "d194_peak", "m194_off", "m194_peak",
            "d195_so", "d195_off", "d195_peak", "m195_so", "m195_off", "m195_peak"]
EFF_LABELS = {
    "basic_day": "Basic service ($/day)", "der_rate": "DER ($/kWh)", "mfc_rate": "MFC ($/kWh)",
    "cbc_rate": "CBC ($/kW/day)", "dsa_rate": "Delivery Svc Adj ($/kWh)",
    "rda_rate": "Rev Decoupling Adj ($/kWh)", "nysa_pct": "NY State Assmt (frac)",
    "pilots_pct": "PILOTs (frac)", "suffolk_pct": "Property Tax Adj (frac)",
    "salestax_pct": "Sales tax (frac)", "psc": "Supply rate ($/kWh)",
    "d580_t1": "580 tier1 ($/kWh)", "d580_t2": "580 tier2 ($/kWh)", "d580_exc": "580 excess ($/kWh)",
    "d180_t1": "180 first250 ($/kWh)", "d180_exc": "180 excess ($/kWh)",
    "d194_off": "194 off-peak ($/kWh)", "d194_peak": "194 peak ($/kWh)",
    "m194_off": "194 off PSC x", "m194_peak": "194 peak PSC x",
    "d195_so": "195 super-off ($/kWh)", "d195_off": "195 off-peak ($/kWh)",
    "d195_peak": "195 peak ($/kWh)", "m195_so": "195 super-off PSC x",
    "m195_off": "195 off PSC x", "m195_peak": "195 peak PSC x",
}


def write_rates_tab(wb, sched):
    """Write the 'Rates' tab: the raw monthly schedule the analysis is built from.

    One row per month showing the season-appropriate delivery tiers, supply PSC, and
    riders, with a provenance flag. Estimated/stale/carried months are shaded yellow
    so the user can see which figures to trust."""
    ws = wb.create_sheet("Rates")
    ws.cell(1, 1, "PSEG LI MONTHLY RATE SCHEDULE (source)").font = Font(FONT, bold=True, size=13)
    ws.cell(2, 1, "Blue = inputs. 'flag' shows data provenance. Dashboard uses date-weighted "
                  "blends of these per billing cycle.").font = Font(FONT, italic=True, size=9, color="5F5E5A")
    hdr = ["Month", "flag", "Supply PSC", "Basic/day", "DER", "MFC", "CBC",
           "580 t1", "580 t2", "580 exc", "180 first", "180 exc",
           "194 off", "194 peak", "195 so", "195 off", "195 peak"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(4, j, h); c.font = Font(FONT, bold=True, color=WHITE, size=9); c.fill = HEAD
    r = 5
    for mk in sorted(sched):
        rec = sched[mk]; s = season_of(int(mk[5:7]))   # show each month's own season's rates
        d = rec["delivery"]
        vals = [mk, rec.get("flag", ""), rec["psc"], rec["basic_day"], rec["der_rate"],
                rec["mfc_rate"], rec["cbc_rate"],
                d["580"][f"t1_{s}"], d["580"][f"t2_{s}"], d["580"][f"exc_{s}"],
                d["180"][f"t1_{s}"], d["180"][f"exc_{s}"],
                d["194"][f"off_{s}"], d["194"][f"peak_{s}"],
                d["195"][f"so_{s}"], d["195"][f"off_{s}"], d["195"][f"peak_{s}"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v)
            c.font = Font(FONT, size=9, color=BLUE if j >= 3 else BLACK)   # numeric inputs in blue
            if j >= 3:
                c.number_format = RFMT
            if rec.get("flag") in ("estimate", "stale", "carried") and j >= 3:
                c.fill = YEL                               # flag unverified numbers in yellow
        r += 1
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 10
    for col in range(3, 18):
        ws.column_dimensions[get_column_letter(col)].width = 9
    ws.freeze_panes = "C5"


def write_effective_tab(wb, prims):
    """Write the 'EffectiveRates' tab and return a cell-reference map for the Dashboard.

    Each column is one billing cycle; each row is a rate parameter, holding the
    day-weighted blend computed in effective_rates. The returned dict maps
    (cycle_index, parameter_key) -> an absolute Excel reference, so the Dashboard's
    formulas can point at these cells and stay live when a user edits them."""
    ws = wb.create_sheet("EffectiveRates")
    ws.cell(1, 1, "EFFECTIVE RATES PER CYCLE (date-weighted from Rates schedule)").font = \
        Font(FONT, bold=True, size=12)
    ws.cell(2, 1, "Each value is the day-weighted average of the monthly schedule across the "
                  "cycle - blends rate changes & seasons.").font = Font(FONT, italic=True, size=9, color="5F5E5A")
    ws.cell(4, 1, "Parameter").font = Font(FONT, bold=True, color=WHITE, size=9)
    ws.cell(4, 1).fill = HEAD
    for i, p in enumerate(prims):
        c = ws.cell(4, 2 + i, f"Cycle {i+1}"); c.font = Font(FONT, bold=True, color=WHITE, size=9)
        c.fill = HEAD; c.alignment = Alignment(horizontal="center")
    eref = {}
    for ri, key in enumerate(EFF_KEYS):
        rr = 5 + ri
        ws.cell(rr, 1, EFF_LABELS[key]).font = Font(FONT, size=9)
        for i, p in enumerate(prims):
            c = ws.cell(rr, 2 + i, round(p["eff"][key], 6))
            c.font = Font(FONT, color=BLUE, size=9); c.number_format = RFMT
            # Record the absolute address so Dashboard formulas can reference this cell.
            eref[(i, key)] = f"'EffectiveRates'!${get_column_letter(2 + i)}${rr}"
    ws.column_dimensions["A"].width = 24
    for i in range(len(prims)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 11
    ws.freeze_panes = "B5"
    return eref


def build_dashboard(wb, prims, eref, meta):
    """Build the 'Dashboard' tab: per-cycle usage, billable-after-banking quantities,
    a full line-item bill for each of the four rate plans, and a lowest-cost comparison.

    All money cells are live Excel formulas that reference the EffectiveRates tab via
    `eref`, so edited rates reflow the whole sheet. Layout: column 1 holds labels,
    columns 2..ncyc+1 hold one cycle each, and the final column holds the row total."""
    ws = wb.active; ws.title = "Dashboard"
    ncyc = len(prims)
    cyc_cols = [2 + i for i in range(ncyc)]   # data columns: one per billing cycle
    tot_col = 2 + ncyc                        # final column: per-row total
    L = get_column_letter
    cl = [L(c) for c in cyc_cols]             # column letters for each cycle
    tl = L(tot_col)                           # column letter for the total

    def H(rr, cc, val, fill=None, color=BLACK, bold=False, align="right"):
        """Write a styled header/label cell and return it (local styling shorthand)."""
        c = ws.cell(rr, cc, val); c.font = Font(FONT, bold=bold, color=color, size=10)
        if fill:
            c.fill = fill
        c.alignment = Alignment(horizontal=align)
        return c

    ws.cell(1, 1, "PSEG Long Island - Billing Analysis").font = Font(FONT, bold=True, size=15)
    ws.cell(2, 1, meta).font = Font(FONT, italic=True, size=9, color="5F5E5A")
    row = 4
    H(row, 1, "BILLING CYCLE", HEAD, WHITE, True, "left")
    for i, c in enumerate(cyc_cols):
        H(row, c, f"Cycle {i+1}", HEAD, WHITE, True)
    H(row, tot_col, "TOTAL", HEAD, WHITE, True)
    row += 1
    H(row, 1, "Service period", SUB, align="left")
    for i, c in enumerate(cyc_cols):
        H(row, c, f"{prims[i]['start']:%m/%d/%y}-{prims[i]['end_label']:%m/%d/%y}", SUB, align="center")
    H(row, tot_col, "All cycles", SUB, align="center")
    row += 1

    ws.cell(row, 1, "USAGE").font = Font(FONT, bold=True, color="0F6E56", size=11)
    row += 1
    urow = {}

    def usage(label, key, agg="sum"):
        """Write one USAGE row of per-cycle quantities from prims[i][key], plus a
        SUM (or MAX, for demand) total, and record its row number in `urow` so the
        rate blocks can reference these cells."""
        nonlocal row
        ws.cell(row, 1, label).font = Font(FONT, size=10)
        for i, c in enumerate(cyc_cols):
            cc = ws.cell(row, c, prims[i][key]); cc.font = Font(FONT, color=BLUE, size=10)
            cc.number_format = KWH; cc.alignment = Alignment(horizontal="right")
        rng = f"{cl[0]}{row}:{cl[-1]}{row}"
        # Demand totals as a MAX (kW doesn't sum across cycles); everything else sums.
        t = ws.cell(row, tot_col, f"=SUM({rng})" if agg == "sum" else f"=MAX({rng})")
        t.font = Font(FONT, bold=True, size=10); t.number_format = KWH
        urow[key] = row; row += 1

    usage("Days", "days"); usage("Net consumption (kWh)", "net")
    usage("  Super-off-peak 10p-6a (kWh)", "so"); usage("  Off-peak (195 basis) (kWh)", "off195")
    usage("  Off-peak (194 basis) (kWh)", "off194"); usage("  On-peak wkdy 3-7p (kWh)", "peak")
    usage("Peak demand (kW)", "demand", "max")

    ws.cell(row, 1, "BILLABLE AFTER CARRY-FORWARD BANKING (what each plan is billed on)").font = \
        Font(FONT, bold=True, color="0F6E56", size=11)
    row += 1
    usage("  Net - flat/tiered 580/180 (kWh)", "b_flat")
    usage("  Peak - 194 (kWh)", "b194_peak"); usage("  Off-peak - 194 (kWh)", "b194_off")
    usage("  Super-off - 195 (kWh)", "b195_so"); usage("  Off-peak - 195 (kWh)", "b195_off")
    usage("  Peak - 195 (kWh)", "b195_peak")

    def u(i, key):
        """Return the A1 reference to cycle i's USAGE cell for `key` (e.g. 'C9')."""
        return f"{cl[i]}{urow[key]}"

    rate_total_rows = {}

    def line(label, fs, bold=False, fill=None):
        """Write one bill line: a per-cycle formula list `fs` plus a SUM total.
        Returns the row number so later formulas (subtotals/totals) can sum it."""
        nonlocal row
        lc = ws.cell(row, 1, label); lc.font = Font(FONT, size=10, bold=bold)
        if fill:
            lc.fill = fill
        for i, c in enumerate(cyc_cols):
            cc = ws.cell(row, c, fs[i]); cc.font = Font(FONT, color=BLACK, size=10, bold=bold)
            cc.number_format = MONEY
            if fill:
                cc.fill = fill
        t = ws.cell(row, tot_col, f"=SUM({cl[0]}{row}:{cl[-1]}{row})")
        t.font = Font(FONT, bold=True, size=10); t.number_format = MONEY
        if fill:
            t.fill = fill
        this = row; row += 1; return this

    def block(rate):
        """Emit the full line-item bill for one rate plan (580/180/194/195).

        Every plan is billed on the post-banking quantities (the BILLABLE AFTER
        BANKING rows), so the comparison is consistent. Delivery and supply differ
        by plan (tiered for 580/180, per-period for the TOD rates); the remaining
        lines (basic, CBC, MFC, DER/DSA/RDA, riders, sales tax) share one shape."""
        nonlocal row
        row += 1
        t = ws.cell(row, 1, RATE_TITLES[rate]); t.font = Font(FONT, bold=True, color=WHITE, size=10)
        for c in cyc_cols + [tot_col]:
            ws.cell(row, c).fill = HEAD
        t.fill = HEAD; row += 1

        def e(i, key):
            """A1 reference to cycle i's EffectiveRates cell for `key` (live rate input)."""
            return eref[(i, key)]

        # Build per-cycle delivery & supply formulas, and the net-billed kWh each plan
        # is charged riders/adjustments on. All reference the banked (b_*) quantities.
        deliv, supply, netb = [], [], []
        for i in range(ncyc):
            days = u(i, "days")
            if rate == "580":
                # Tiered: first 250 kWh/30d at t1, next 150 at t2, remainder at excess.
                nb = u(i, "b_flat"); netb.append(nb)
                t1, t2 = f"(250*{days}/30)", f"(150*{days}/30)"
                deliv.append(f"=MIN({nb},{t1})*{e(i,'d580_t1')}"
                             f"+MIN(MAX({nb}-{t1},0),{t2})*{e(i,'d580_t2')}"
                             f"+MAX({nb}-{t1}-{t2},0)*{e(i,'d580_exc')}")
                supply.append(f"={nb}*{e(i,'psc')}")
            elif rate == "180":
                # Two-tier: first 250 kWh/30d at t1, remainder at excess.
                nb = u(i, "b_flat"); netb.append(nb)
                t1 = f"(250*{days}/30)"
                deliv.append(f"=MIN({nb},{t1})*{e(i,'d180_t1')}+MAX({nb}-{t1},0)*{e(i,'d180_exc')}")
                supply.append(f"={nb}*{e(i,'psc')}")
            elif rate == "194":
                # Two TOU periods; supply = PSC scaled by each period's multiplier.
                pk, of = u(i, "b194_peak"), u(i, "b194_off"); netb.append(f"({pk}+{of})")
                deliv.append(f"={pk}*{e(i,'d194_peak')}+{of}*{e(i,'d194_off')}")
                supply.append(f"={e(i,'psc')}*({of}*{e(i,'m194_off')}+{pk}*{e(i,'m194_peak')})")
            else:
                # Three TOU periods (super-off / off / peak); same supply-multiplier scheme.
                so, of, pk = u(i, "b195_so"), u(i, "b195_off"), u(i, "b195_peak")
                netb.append(f"({so}+{of}+{pk})")
                deliv.append(f"={so}*{e(i,'d195_so')}+{of}*{e(i,'d195_off')}+{pk}*{e(i,'d195_peak')}")
                supply.append(f"={e(i,'psc')}*({so}*{e(i,'m195_so')}"
                              f"+{of}*{e(i,'m195_off')}+{pk}*{e(i,'m195_peak')})")

        br = line("Basic Service", [f"={e(i,'basic_day')}*{u(i,'days')}" for i in range(ncyc)])
        cr = line("Customer Benefit Contribution",
                  [f"={e(i,'cbc_rate')}*{u(i,'demand')}*{u(i,'days')}" for i in range(ncyc)])
        mr = line("Merchant Function Charge", [f"={e(i,'mfc_rate')}*{netb[i]}" for i in range(ncyc)])
        dr = line("Delivery (energy)", deliv)
        dsr = line("Delivery & System subtotal",
                   [f"={cl[i]}{br}+{cl[i]}{cr}+{cl[i]}{mr}+{cl[i]}{dr}" for i in range(ncyc)],
                   bold=True, fill=SUB)
        sr = line("Power Supply", supply)
        der = line("DER Charge", [f"={e(i,'der_rate')}*{netb[i]}" for i in range(ncyc)])
        dsa = line("Delivery Service Adjustment", [f"={e(i,'dsa_rate')}*{netb[i]}" for i in range(ncyc)])
        rda = line("Revenue Decoupling Adjustment", [f"={e(i,'rda_rate')}*{netb[i]}" for i in range(ncyc)])
        # Riders (NYSA, PILOTs, property tax) are a percentage of the delivery+supply base.
        base = [f"({cl[i]}{dsr}+{cl[i]}{sr})" for i in range(ncyc)]
        nr = line("NY State Assessment", [f"={e(i,'nysa_pct')}*{base[i]}" for i in range(ncyc)])
        pr = line("Revenue-Based PILOTs", [f"={e(i,'pilots_pct')}*{base[i]}" for i in range(ncyc)])
        sf = line("Property Tax Adj", [f"={e(i,'suffolk_pct')}*{base[i]}" for i in range(ncyc)])
        # Pre-tax = everything above; sales tax then applies to that whole pre-tax sum.
        pretax = [f"({cl[i]}{dsr}+{cl[i]}{sr}+{cl[i]}{der}+{cl[i]}{dsa}+{cl[i]}{rda}"
                  f"+{cl[i]}{nr}+{cl[i]}{pr}+{cl[i]}{sf})" for i in range(ncyc)]
        sx = line("Sales Tax", [f"={e(i,'salestax_pct')}*{pretax[i]}" for i in range(ncyc)])
        tr = line("TOTAL CHARGES", [f"={pretax[i]}+{cl[i]}{sx}" for i in range(ncyc)], bold=True, fill=TOT)
        rate_total_rows[rate] = tr   # remember this plan's TOTAL row for the comparison block

    for rate in ["580", "180", "194", "195"]:
        block(rate)

    # Side-by-side comparison of the four plans' totals, plus the cheapest per cycle.
    row += 2
    ws.cell(row, 1, "RATE COMPARISON - total bill ($)").font = Font(FONT, bold=True, color="0F6E56", size=11)
    row += 1
    H(row, 1, "Rate plan", HEAD, WHITE, True, "left")
    for i, c in enumerate(cyc_cols):
        H(row, c, f"Cycle {i+1}", HEAD, WHITE, True)
    H(row, tot_col, "TOTAL", HEAD, WHITE, True)
    row += 1
    cmp = {}
    for rate in ["580", "180", "194", "195"]:
        ws.cell(row, 1, RATE_TITLES[rate]).font = Font(FONT, size=10)
        tr = rate_total_rows[rate]
        # Mirror each plan's TOTAL CHARGES row up into the comparison table.
        for i, c in enumerate(cyc_cols):
            cc = ws.cell(row, c, f"={cl[i]}{tr}"); cc.font = Font(FONT, color=GREEN, size=10)
            cc.number_format = MONEY
        t = ws.cell(row, tot_col, f"={tl}{tr}"); t.font = Font(FONT, bold=True, size=10); t.number_format = MONEY
        cmp[rate] = row; row += 1
    ws.cell(row, 1, "Lowest-cost plan").font = Font(FONT, bold=True, size=10)
    # The four plan rows are contiguous (580..195); MATCH(MIN(range)) finds the cheapest,
    # and INDEX maps that position back to the plan name.
    f0, f1 = cmp["580"], cmp["195"]
    for i, c in enumerate(cyc_cols):
        rng = f"{cl[i]}{f0}:{cl[i]}{f1}"
        cc = ws.cell(row, c, f'=INDEX({{"580";"180";"194";"195"}},MATCH(MIN({rng}),{rng},0))')
        cc.font = Font(FONT, bold=True, color="0F6E56", size=10)
        cc.alignment = Alignment(horizontal="center"); cc.fill = SUB
    rng = f"{tl}{f0}:{tl}{f1}"   # same logic for the all-cycles TOTAL column
    t = ws.cell(row, tot_col, f'=INDEX({{"580";"180";"194";"195"}},MATCH(MIN({rng}),{rng},0))')
    t.font = Font(FONT, bold=True, color="0F6E56", size=10)
    t.alignment = Alignment(horizontal="center"); t.fill = SUB
    row += 2
    ws.cell(row, 1, "Date-matched estimate: each cycle uses day-weighted rates from the Rates schedule "
                    "(blends rate changes & seasons). All four plans are billed on kWh AFTER carry-forward "
                    "banking (see the BILLABLE AFTER BANKING rows), so the comparison is consistent; the "
                    "Time-of-Day plans use the Banking tab's bank logic, not optimistic in-cycle netting. "
                    "PILOT/property-tax recovery is LIPA-wide; only local sales tax varies by county. "
                    "Yellow cells on Rates = estimated/stale; refresh with --refresh while online.").font = \
        Font(FONT, italic=True, size=8, color="5F5E5A")

    ws.column_dimensions["A"].width = 32
    for c in cyc_cols + [tot_col]:
        ws.column_dimensions[L(c)].width = 13
    ws.freeze_panes = "B5"


# --------------------------------------------------------------------------- #
# Energy-bank simulation (Time-of-Day rates 194 and 195)                       #
# --------------------------------------------------------------------------- #
# PSEG LI TOD credit banks exchange at fixed value ratios per the Excess
# Generation Exchange form: Rate 195 uses three banks at 1:2:4
# (peak:off:super-off); Rate 194 uses two banks at 1:2 (peak:off).
TOD_PERIODS: "dict[str, tuple[str, ...]]" = {"194": ("peak", "off"), "195": ("peak", "off", "so")}
TOD_WEIGHTS: "dict[str, dict[str, float]]" = {
    "194": {"peak": 2.0, "off": 1.0},
    "195": {"peak": 4.0, "off": 2.0, "so": 1.0},
}
# Which key in `prims` holds the per-cycle net for each (rate, period).
TOD_NETKEY: "dict[str, dict[str, str]]" = {
    "194": {"peak": "peak", "off": "off194"},
    "195": {"peak": "peak", "off": "off195", "so": "so"},
}
PERIOD_LABEL = {"peak": "PEAK", "off": "OFF", "so": "SO"}
MAX_TRANSFERS_PER_CYCLE = 2  # the form allows up to two statements per request


def _xfer_increment(weights: "dict[str, float]", src: str, dst: str) -> int:
    """Whole-kWh increment required on the SOURCE bank (per the form): 1 when moving to
    a lower/equal-value bank, else the integer value ratio (2 or 4)."""
    return max(1, round(weights[dst] / weights[src]))


def _cover_deficits(banks: "dict[str, float]", billable: "dict[str, float]",
                    weights: "dict[str, float]", max_transfers: int
                    ) -> "list[tuple[str, str, float, float]]":
    """Profile-agnostic greedy: move carried-in bank surplus to cover this cycle's
    remaining per-period deficits, using value-neutral transfers in whole-kWh source
    increments, at most `max_transfers` moves. Works for any usage profile and for
    either rate (two or three banks). Deficits are covered costliest-first and drained
    from the lowest-value surplus bank first, preserving the most flexible credits and
    avoiding mildly-lossy directions unless they are the only option."""
    periods = tuple(weights)
    transfers: "list[tuple[str, str, float, float]]" = []
    for dst in sorted((d for d in periods if billable[d] > 1e-9), key=lambda d: -weights[d]):
        for src in sorted((s for s in periods if banks[s] > 1e-9 and s != dst),
                          key=lambda s: weights[s]):
            if billable[dst] <= 1e-9 or len(transfers) >= max_transfers:
                break
            inc = _xfer_increment(weights, src, dst)
            gain = weights[src] / weights[dst]            # destination kWh per source kWh
            by_bank = math.floor(banks[src] / inc) * inc
            by_need = math.floor((billable[dst] / gain) / inc) * inc
            src_kwh = float(min(by_bank, by_need))
            if src_kwh > 0:
                dst_kwh = src_kwh * gain
                banks[src] -= src_kwh; billable[dst] -= dst_kwh
                transfers.append((src, dst, src_kwh, dst_kwh))
        if len(transfers) >= max_transfers:
            break
    return transfers


def simulate_banks(prims: "list[dict[str, Any]]", rate: str
                   ) -> "tuple[list[dict[str, Any]], dict[str, float]]":
    """Sequential per-period kWh credit banks with carry-forward and the greedy transfer
    step, for Rate 194 (two banks) or 195 (three banks).

    Per cycle, per period: net = consumption - generation (already in prims).
      net  < 0 -> surplus deposited to that period's bank (available next cycle).
      net >= 0 -> drawn down against that period's own carried-in bank first.
    Remaining deficits are covered by transferring carried-in surplus between banks at
    the form's value ratios, whole-kWh, at most two per cycle. Deposits made this cycle
    are added afterward, so they are only usable from the next cycle."""
    periods = TOD_PERIODS[rate]
    weights = TOD_WEIGHTS[rate]
    nk = TOD_NETKEY[rate]
    banks: "dict[str, float]" = {k: 0.0 for k in periods}
    rows: "list[dict[str, Any]]" = []
    for p in prims:
        bank_in = dict(banks)
        net = {k: float(p[nk[k]]) for k in periods}
        billable: "dict[str, float]" = {}
        withdraw: "dict[str, float]" = {}
        pending: "dict[str, float]" = {}
        for k in periods:
            if net[k] >= 0:
                w = min(net[k], banks[k]); banks[k] -= w
                withdraw[k] = w; billable[k] = net[k] - w; pending[k] = 0.0
            else:
                withdraw[k] = 0.0; billable[k] = 0.0; pending[k] = -net[k]
        transfers = _cover_deficits(banks, billable, weights, MAX_TRANSFERS_PER_CYCLE)
        for k in periods:
            banks[k] += pending[k]
        rows.append(dict(start=p["start"], end_label=p.get("end_label", p["end"]),
                         days=int(p["days"]), demand=float(p["demand"]), eff=p["eff"],
                         periods=periods, net=net, billable=dict(billable),
                         deposit=dict(pending), withdraw=dict(withdraw),
                         transfers=transfers, bank_in=bank_in, bank_out=dict(banks)))
    return rows, banks


def bill_tou(rate: str, billable: "dict[str, float]", net_billed: float, days: int,
             demand: float, e: "dict[str, Any]") -> float:
    """All-in charge for one cycle on Rate 194 or 195 from billable kWh per period.
    Mirrors the Dashboard line items (basic, CBC, MFC, delivery, supply, DER/DSA/RDA,
    NYSA/PILOTs/property tax, sales tax)."""
    if rate == "195":
        deliv = (billable["so"] * e["d195_so"] + billable["off"] * e["d195_off"]
                 + billable["peak"] * e["d195_peak"])
        supply = e["psc"] * (billable["so"] * e["m195_so"] + billable["off"] * e["m195_off"]
                             + billable["peak"] * e["m195_peak"])
    else:  # 194
        deliv = billable["off"] * e["d194_off"] + billable["peak"] * e["d194_peak"]
        supply = e["psc"] * (billable["off"] * e["m194_off"] + billable["peak"] * e["m194_peak"])
    ds_sub = e["basic_day"] * days + e["cbc_rate"] * demand * days + e["mfc_rate"] * net_billed + deliv
    der = e["der_rate"] * net_billed
    dsa = e["dsa_rate"] * net_billed
    rda = e["rda_rate"] * net_billed
    base = ds_sub + supply
    riders = (e["nysa_pct"] + e["pilots_pct"] + e["suffolk_pct"]) * base
    pretax = ds_sub + supply + der + dsa + rda + riders
    return pretax * (1.0 + e["salestax_pct"])


def value_pool_optimum(prims: "list[dict[str, Any]]", rate: str) -> float:
    """Frictionless lower bound: treat all credits as one pool of value units, carry it
    forward one cycle, and cover each cycle's deficits from the pool. Because the
    transfer ratios equal the value weights, ordering is immaterial; the gap to the
    greedy simulation is the total cost of real transfer frictions."""
    periods = TOD_PERIODS[rate]
    weights = TOD_WEIGHTS[rate]
    nk = TOD_NETKEY[rate]
    pool = 0.0
    total = 0.0
    for p in prims:
        net = {k: float(p[nk[k]]) for k in periods}
        billable = {k: max(net[k], 0.0) for k in periods}
        for k in sorted(periods, key=lambda k: -weights[k]):
            if pool <= 0:
                break
            cover = min(billable[k], pool / weights[k])
            billable[k] -= cover; pool -= cover * weights[k]
        pool += sum(-net[k] * weights[k] for k in periods if net[k] < 0)
        total += bill_tou(rate, billable, sum(billable.values()), int(p["days"]),
                          float(p["demand"]), p["eff"])
    return total


def profile_summary(prims: "list[dict[str, Any]]", rate: str) -> "dict[str, Any]":
    """Classify the household's annual net per period (consumer vs producer) and flag
    forfeiture exposure. value_net < 0 means more credit value is banked than consumed,
    so some credits never find a deficit and accumulate toward the 20-year forfeiture."""
    periods = TOD_PERIODS[rate]
    weights = TOD_WEIGHTS[rate]
    nk = TOD_NETKEY[rate]
    annual = {k: sum(float(p[nk[k]]) for p in prims) for k in periods}
    value_net = sum(annual[k] * weights[k] for k in periods)
    return dict(annual=annual, value_net=value_net, net_producer=value_net < 0)


def banked_total(prims: "list[dict[str, Any]]", rate: str) -> float:
    """Total all-in cost for `rate` across all cycles using carry-forward banking."""
    sim, _ = simulate_banks(prims, rate)
    return sum(bill_tou(rate, s["billable"], sum(s["billable"].values()),
                        s["days"], s["demand"], s["eff"]) for s in sim)


def attach_banked_quantities(prims: "list[dict[str, Any]]") -> None:
    """Annotate each cycle with the kWh it is actually billed on after carry-forward
    banking, so the Dashboard compares every plan on the same realistic basis:
      b_flat                 - net after a single volumetric bank (Rates 580/180);
                               equals metered net whenever the cycle is a net import.
      b194_peak / b194_off   - billable per period after Rate 194 banking (two banks).
      b195_so/off/peak       - billable per period after Rate 195 banking (three banks).
    Banking decisions depend only on kWh and the fixed value ratios, not on the dollar
    rates, so these quantities stay valid when rate cells are edited on the Rates tab."""
    bank = 0.0
    for p in prims:
        net = float(p["net"])
        if net >= 0:
            applied = min(net, bank); bank -= applied; p["b_flat"] = net - applied
        else:
            p["b_flat"] = 0.0; bank += -net
    for rate, pref in (("194", "b194"), ("195", "b195")):
        sim, _ = simulate_banks(prims, rate)
        for p, s in zip(prims, sim):
            for k in TOD_PERIODS[rate]:
                p[f"{pref}_{k}"] = float(s["billable"][k])


def write_banking_tab(wb: "Workbook", sim: "list[dict[str, Any]]",
                      leftover: "dict[str, float]", optimum: float,
                      profile: "dict[str, Any]", rate: str) -> float:
    """Write the 'Banking194'/'Banking195' detail tab and return the banked total.

    Columns adapt to the rate's bank count (two for 194, three for 195). Per cycle it
    shows net by period, billable-after-banking by period, the transfers made, and the
    ending bank balances, then the all-in charge. Below the table it adds an optimality
    check (greedy vs frictionless value-pool bound) and a usage-profile/forfeiture
    block. `leftover`, `optimum`, and `profile` are precomputed by the caller."""
    periods = TOD_PERIODS[rate]
    up = PERIOD_LABEL
    ratio = ":".join(str(int(w)) for w in sorted(TOD_WEIGHTS[rate].values()))   # e.g. 1:2 or 1:2:4
    names = "peak:off:super-off" if rate == "195" else "peak:off"
    ws = wb.create_sheet(f"Banking{rate}")
    t = ws.cell(1, 1, f"RATE {rate} ENERGY-BANK SIMULATION  ({ratio} {names} transfers)")
    t.font = Font(FONT, bold=True, color="0F6E56", size=11)
    ws.cell(2, 1, "Per-cycle kWh: net by period, bank balances, credit transfers, "
                  f"billable after banking, and the resulting all-in {rate} charge.").font = \
        Font(FONT, italic=True, size=9, color="5F5E5A")
    hdr = (["Service period", "days"]
           + [f"{up[k]} net" for k in periods]
           + [f"{up[k]} bill" for k in periods]
           + ["transfers (src\u2192dst: src kWh)"]
           + [f"{up[k]} bank" for k in periods]
           + [f"{rate} all-in $"])
    xfer_col = 2 + 2 * len(periods) + 1
    r = 4
    for c, h in enumerate(hdr, start=1):
        cell = ws.cell(r, c, h); cell.font = Font(FONT, bold=True, color=WHITE, size=9)
        cell.fill = HEAD; cell.alignment = Alignment(horizontal="center")
    r += 1
    total = 0.0
    for s in sim:
        e = s["eff"]; b = s["billable"]
        net_billed = sum(b[k] for k in periods)
        amt = bill_tou(rate, b, net_billed, s["days"], s["demand"], e)
        total += amt
        xtext = "; ".join(f"{up[src]}\u2192{up[dst]}: {sk:.0f}"
                          for src, dst, sk, _g in s["transfers"]) or "\u2014"
        vals: "list[Any]" = ([f"{s['start']:%m/%d/%y}-{s['end_label']:%m/%d/%y}", s["days"]]
                             + [round(s["net"][k]) for k in periods]
                             + [round(b[k]) for k in periods]
                             + [xtext]
                             + [round(s["bank_out"][k]) for k in periods]
                             + [amt])
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(r, c, v); cell.font = Font(FONT, size=9)
            if c == len(vals):
                cell.number_format = MONEY
            elif c not in (1, xfer_col):
                cell.number_format = KWH
            if c == xfer_col:
                cell.alignment = Alignment(horizontal="left")
        r += 1
    tc = ws.cell(r, 1, "TOTAL (simulated cycles)"); tc.font = Font(FONT, bold=True, size=9)
    for c in range(1, len(hdr) + 1):
        ws.cell(r, c).fill = TOT
    tot = ws.cell(r, len(hdr), total); tot.font = Font(FONT, bold=True, size=9); tot.number_format = MONEY
    r += 2

    oc = ws.cell(r, 1, "OPTIMALITY CHECK"); oc.font = Font(FONT, bold=True, color="0F6E56", size=10); r += 1
    ws.cell(r, 1, "Greedy strategy (this tab)").font = Font(FONT, size=9)
    gv = ws.cell(r, len(hdr), total); gv.number_format = MONEY; gv.font = Font(FONT, size=9); r += 1
    ws.cell(r, 1, "Frictionless value-pool optimum").font = Font(FONT, size=9)
    ov = ws.cell(r, len(hdr), optimum); ov.number_format = MONEY; ov.font = Font(FONT, size=9); r += 1
    ws.cell(r, 1, "Gap (cost of transfer frictions)").font = Font(FONT, bold=True, size=9)
    gp = ws.cell(r, len(hdr), total - optimum); gp.number_format = MONEY; gp.font = Font(FONT, bold=True, size=9)
    r += 2

    pc = ws.cell(r, 1, "USAGE PROFILE (annual net by period)")
    pc.font = Font(FONT, bold=True, color="0F6E56", size=10); r += 1
    a = profile["annual"]
    for k in periods:
        lab = "net consumer (banks drain)" if a[k] >= 0 else "net producer (banks build)"
        ws.cell(r, 1, f"{up[k]}: {a[k]:+,.0f} kWh/yr  \u2014 {lab}").font = Font(FONT, size=9)
        r += 1
    if profile["net_producer"]:
        msg = ("Household is a NET PRODUCER in value terms: some credits have no deficit to offset and "
               "would accumulate toward the 20-year forfeiture. Consider a larger load shift or system review.")
    else:
        msg = ("Household is a net consumer in value terms: credits are fully absorbed by usage, so nothing "
               "is exposed to the 20-year forfeiture.")
    ws.cell(r, 1, msg).font = Font(FONT, italic=True, size=9, color="5F5E5A")
    r += 2

    lo = "  ".join(f"{up[k].lower()} {leftover[k]:.0f}" for k in periods)
    ws.cell(r, 1, f"Bank balance at end of window: {lo} kWh.  Rolls forward; per the PSEG "
                  f"operation letter, residential (non-demand) credits carry over for up to 20 years "
                  f"(only any balance still unused after year 20 is forfeited) -- no annual buyback.").font = \
        Font(FONT, italic=True, size=9, color="5F5E5A")
    ws.column_dimensions["A"].width = 24
    for c in range(2, len(hdr) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 26 if c == xfer_col else 11
    return total


# --------------------------------------------------------------------------- #
def main():
    """CLI entry point: parse args, load usage and rates, build the workbook.

    Pipeline: load the interval CSV -> ensure/refresh the rate schedule -> apply the
    county sales tax -> split into billing cycles from --read-dates -> reduce to
    per-cycle primitives -> compute billable-after-banking quantities -> write the
    Dashboard, Rates, EffectiveRates, and both Banking tabs -> save."""
    ap = argparse.ArgumentParser(description="Analyze PSEG LI consumption into an xlsx dashboard.")
    ap.add_argument("--cons_file", required=True)
    ap.add_argument("--read-dates", dest="read_dates", nargs="*", default=[],
                    help="Meter-read boundary dates YYYY-MM-DD ... (each bill's Service From "
                         "plus the last bill's Service To). N+1 dates -> N half-open cycles "
                         "[read_i, read_i+1). If omitted, the whole period is one cycle.")
    ap.add_argument("--out", default=None)
    ap.add_argument("--demand", type=float, default=None, help="Override peak demand (kW) for CBC")
    cty = ap.add_mutually_exclusive_group()
    cty.add_argument("--suffolk", dest="county", action="store_const", const="suffolk",
                     help="Suffolk County tax treatment (default).")
    cty.add_argument("--nassau", dest="county", action="store_const", const="nassau",
                     help="Nassau County tax treatment (residential energy is sales-tax exempt).")
    cty.add_argument("--rockaways", dest="county", action="store_const", const="rockaways",
                     help="Rockaways / NYC tax treatment.")
    ap.set_defaults(county="suffolk")
    ap.add_argument("--sales-tax", dest="sales_tax", type=float, default=None,
                    help="Override the local sales-tax fraction (e.g. 0.03) regardless of county.")
    ap.add_argument("--refresh", action="store_true", help="Force re-scrape of estimated months")
    args = ap.parse_args()

    if not os.path.exists(args.cons_file):
        sys.exit(f"File not found: {args.cons_file}")
    try:
        # Accept dates as separate args or as one space-joined string; dedupe and sort.
        rd_tokens = [tok for chunk in args.read_dates for tok in str(chunk).split()]
        read_dates = sorted({dt.datetime.strptime(s, "%Y-%m-%d").date() for s in rd_tokens})
    except ValueError as e:
        sys.exit(f"Read dates must be YYYY-MM-DD (one per date): {e}")
    if read_dates and len(read_dates) < 2:
        # N cycles need N+1 boundaries, so a lone date can't define a cycle.
        sys.exit("--read-dates needs at least 2 dates (N+1 dates define N cycles).")

    print("Loading consumption data...")
    m = load_consumption(args.cons_file)
    didx = cast("pd.DatetimeIndex", m.index)
    start = cast("pd.Timestamp", didx.min()).date()
    end = cast("pd.Timestamp", didx.max()).date()
    print(f"  {len(m):,} intervals, {start} to {end}")

    print("Loading & refreshing rate schedule...")
    cache = load_cache()
    sched_start = min(read_dates) if read_dates else start
    sched_end = max(read_dates) if read_dates else end
    cache = refresh_schedule(cache, sched_start, sched_end, force=args.refresh)
    sched = cache["schedule"]
    sales_tax = apply_county_taxes(sched, args.county, args.sales_tax)
    note = {"suffolk": "bill-verified", "nassau": "residential energy exempt; Glen Cove/Long Beach SD ~3%",
            "rockaways": "NYC rate, estimated -- confirm against your bill"}[args.county]
    src = "override" if args.sales_tax is not None else note
    print(f"  Service area: {args.county.capitalize()} -- local sales tax {sales_tax:.2%} ({src}).")

    warnings = set()
    if read_dates:
        # Half-open [read_i, read_i+1): map to the inclusive-end machinery by
        # using read_i+1 - 1 day as the inclusive end (preserves day count).
        cycles = [(read_dates[i], read_dates[i + 1] - dt.timedelta(days=1))
                  for i in range(len(read_dates) - 1)]
    else:
        cycles = [(start, end)]
    prims = cycle_primitives(m, cycles, federal_holidays(range(sched_start.year, sched_end.year + 1)), sched, warnings)
    if read_dates:
        for p in prims:
            p["end_label"] = p["end"] + dt.timedelta(days=1)
    if args.demand is not None:
        for p in prims:
            p["demand"] = args.demand
    print(f"  {len(prims)} billing cycle(s).")
    for w in sorted(warnings):
        print(f"  WARNING: {w}")

    attach_banked_quantities(prims)
    wb = Workbook()
    write_rates_tab(wb, sched)
    eref = write_effective_tab(wb, prims)
    meta = (f"Source: {os.path.basename(args.cons_file)}  |  Data {start} to {end}  |  "
            f"{len(prims)} cycle(s)  |  Generated {dt.date.today()}")
    build_dashboard(wb, prims, eref, meta)

    # Always write the banking detail for both Time-of-Day rates: the Dashboard
    # already bills them on this bank logic, so the audit trail is never optional.
    t194, t195 = banked_total(prims, "194"), banked_total(prims, "195")
    best = "194" if t194 < t195 else "195"
    print(f"\n  Time-of-Day banking: 194 ${t194:,.2f} vs 195 ${t195:,.2f}  "
          f"->  Rate {best} is the cheaper TOD plan.")
    for rate in ("194", "195"):
        periods = TOD_PERIODS[rate]
        sim, leftover = simulate_banks(prims, rate)
        optimum = value_pool_optimum(prims, rate)
        profile = profile_summary(prims, rate)
        bk_total = write_banking_tab(wb, sim, leftover, optimum, profile, rate)
        a = profile["annual"]
        prof = ", ".join(f"{k.upper()} {a[k]:+,.0f}" for k in periods)
        tag = "NET PRODUCER (forfeiture exposure)" if profile["net_producer"] else "net consumer"
        star = "  <- recommended" if rate == best else ""
        print(f"    Rate {rate}: banked ${bk_total:,.2f}  (optimum ${optimum:,.2f}, "
              f"gap ${bk_total - optimum:.2f}); profile {prof} -> {tag}{star}")

    wb.move_sheet("Dashboard", -(len(wb.sheetnames) - 1))

    out = args.out or os.path.splitext(os.path.basename(args.cons_file))[0] + "_analysis.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
